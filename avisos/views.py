from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import TemplateView, CreateView, UpdateView, FormView, ListView, DeleteView
from django.shortcuts import redirect, get_object_or_404
from django.db.models import Count, Q
from datetime import timedelta

from .forms import SignupForm, UserUpdateForm, StudentForm, NotifyLateForm, SchoolStaffToggleForm, LateArrivalReportFilterForm, LateArrivalAggregatedFilterForm
from .models import User, Student, ResponsibleStudent, LateArrival
from django.contrib.auth import logout
from django.forms import formset_factory


def logout_then_login(request):
    logout(request)
    return redirect("login")

class SchoolOnlyMixin(UserPassesTestMixin):
    def test_func(self):
        u = self.request.user
        return u.is_authenticated and (u.is_superuser or getattr(u, "is_school_staff", False))


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = "home.html"
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["latest_lates"] = LateArrival.objects.filter(responsible=self.request.user).order_by("-reported_at")[:5]
        # Flag escolar para mostrar botón especial
        ctx["is_school_staff"] = self.request.user.is_superuser or getattr(self.request.user, "is_school_staff", False)
        return ctx

class SignupView(CreateView):
    form_class = SignupForm
    def form_valid(self, form):
        resp = super().form_valid(form)
        messages.success(self.request, "¡Cuenta creada! Ya podés usar el sistema.")
        login(self.request, self.object)
        return resp
    template_name = "auth/signup.html"
    success_url = reverse_lazy("home")


class ProfileView(LoginRequiredMixin, UpdateView):
    form_class = UserUpdateForm
    template_name = "avisos/profile.html"
    success_url = reverse_lazy("home")
    def get_object(self):
        return self.request.user

class RegisterStudentsView(LoginRequiredMixin, FormView):
    template_name = "avisos/register_students.html"
    success_url = reverse_lazy("home")
    def get_form_class(self):
        return formset_factory(StudentForm, extra=0)
    def form_valid(self, formset):
        for form in formset:
            if form.cleaned_data:
                student = form.save()
                ResponsibleStudent.objects.get_or_create(responsible=self.request.user, student=student)
        messages.success(self.request, "Alumnos registrados")
        return super().form_valid(formset)



class NotifyLateView(LoginRequiredMixin, FormView):
    template_name = "avisos/notify_late.html"
    form_class = NotifyLateForm
    success_url = reverse_lazy("notifications_list")

    def dispatch(self, request, *args, **kwargs):
        has_students = Student.objects.filter(
            responsiblestudent__responsible=request.user,
            active=True
        ).exists()
        if not has_students:
            messages.warning(request, "No tenés alumnos activos para avisar llegadas tarde.")
            return redirect("home")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs
    
class NotificationsListView(LoginRequiredMixin, ListView):
    model = LateArrival
    template_name = "avisos/notifications_list.html"
    context_object_name = "notifications"
    def get_queryset(self):
        return LateArrival.objects.filter(responsible=self.request.user)


class StudentsListView(LoginRequiredMixin, ListView):
    template_name = "alumnos/list.html"
    context_object_name = "students"
    def get_queryset(self):
        return Student.objects.filter(responsiblestudent__responsible=self.request.user).distinct()

class StudentCreateView(LoginRequiredMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = "alumnos/form.html"
    success_url = reverse_lazy("students_list")
    def form_valid(self, form):
        resp = self.request.user
        obj = form.save()
        ResponsibleStudent.objects.get_or_create(responsible=resp, student=obj)
        messages.success(self.request, "Alumno agregado correctamente.")
        return super().form_valid(form)

class StudentUpdateView(LoginRequiredMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = "alumnos/form.html"
    success_url = reverse_lazy("students_list")
    def get_queryset(self):
        return Student.objects.filter(responsiblestudent__responsible=self.request.user).distinct()
    def form_valid(self, form):
        messages.success(self.request, "Alumno actualizado correctamente.")
        return super().form_valid(form)

class StudentDeleteView(LoginRequiredMixin, DeleteView):
    model = Student
    template_name = "alumnos/confirm_delete.html"
    success_url = reverse_lazy("students_list")
    def get_queryset(self):
        return Student.objects.filter(responsiblestudent__responsible=self.request.user).distinct()
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Alumno eliminado.")
        return super().delete(request, *args, **kwargs)

class LateArrivalDeleteView(LoginRequiredMixin, DeleteView):
    model = LateArrival
    template_name = "avisos/confirm_notification_delete.html"
    success_url = reverse_lazy("notifications_list")

    def get_queryset(self):
        # Solo permite borrar avisos propios
        return LateArrival.objects.filter(responsible=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Aviso eliminado.")
        return super().delete(request, *args, **kwargs)

class SchoolTodayLatesView(LoginRequiredMixin, SchoolOnlyMixin, ListView):
    template_name = "avisos/school_today_lates.html"
    context_object_name = "lates"

    def get_queryset(self):
        today = timezone.localdate()
        return (LateArrival.objects
                .select_related("student", "responsible", "reviewed_by")
                .filter(reported_at__date=today)
                .order_by("-reported_at"))

    def dispatch(self, request, *args, **kwargs):
        today = timezone.localdate()
        qs_hoy_pend = LateArrival.objects.filter(
            reported_at__date=today,
            reviewed_at__isnull=True
        )
        if qs_hoy_pend.exists():
            qs_hoy_pend.update(
                reviewed_by_id=request.user.id,
                reviewed_at=timezone.now()
            )
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            lates = list(ctx["lates"])  # materializamos

            if not lates:
                ctx["lates"] = lates
                return ctx

            last30 = timezone.now() - timedelta(days=30)
            student_ids = {l.student_id for l in lates}

            agg = (LateArrival.objects
                .filter(student_id__in=student_ids)
                .values("student_id")
                .annotate(
                    total=Count("id"),
                    last30=Count("id", filter=Q(reported_at__gte=last30))
                ))

            counts_by_student = {row["student_id"]: row for row in agg}

            # adjuntamos atributos a cada aviso para fácil uso en template
            for l in lates:
                row = counts_by_student.get(l.student_id, {"total": 0, "last30": 0})
                l.total_count = row["total"]
                l.last30_count = row["last30"]

            ctx["lates"] = lates
            return ctx

class StudentLateHistoryView(LoginRequiredMixin, SchoolOnlyMixin, ListView):
    template_name = "avisos/student_late_history.html"
    context_object_name = "lates"

    def get_queryset(self):
        student_id = self.kwargs["student_id"]
        return LateArrival.objects.filter(student_id=student_id).select_related("student","responsible").order_by("-reported_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        student_id = self.kwargs["student_id"]
        ctx["student"] = get_object_or_404(Student, pk=student_id)
        return ctx


# === NUEVO: Asignar personal de escuela (primero solo superuser, luego school_staff también) ===
class SchoolStaffAssignView(LoginRequiredMixin, FormView):
    template_name = "avisos/school_staff_assign.html"
    form_class = SchoolStaffToggleForm
    success_url = reverse_lazy("school_staff_assign")

    def dispatch(self, request, *args, **kwargs):
        any_school = User.objects.filter(is_school_staff=True).exists()
        if not any_school and not request.user.is_superuser:
            messages.error(request, "Solo un superusuario puede asignar el primer personal de escuela.")
            return redirect("home")
        if any_school and not (request.user.is_superuser or request.user.is_school_staff):
            messages.error(request, "No tenés permisos para asignar personal de escuela.")
            return redirect("home")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        id_number = form.cleaned_data["id_number"]
        mark = form.cleaned_data["is_school_staff"]
        user = get_object_or_404(User, id_number=id_number)
        user.is_school_staff = mark
        user.save(update_fields=["is_school_staff"])
        if mark:
            messages.success(self.request, f"{user.full_name} marcado como Personal de escuela.")
        else:
            messages.success(self.request, f"{user.full_name} ya no es Personal de escuela.")
        return super().form_valid(form)
    

def user_is_school(user):
    return user.is_authenticated and (user.is_superuser or getattr(user, "is_school_staff", False))

def scope_late_arrivals_for(user, qs):
    return qs if user_is_school(user) else qs.filter(responsible=user)


from django.utils import timezone
from datetime import timedelta

class LateArrivalReportView(LoginRequiredMixin, TemplateView):
    template_name = "avisos/reports_detailed.html"

    def get(self, request, *args, **kwargs):
        # 1) Construimos los datos del form con defaults (hoy-30 .. hoy)
        data = request.GET.copy()
        today = timezone.localdate()
        default_from = (today - timedelta(days=30)).isoformat()
        default_to = today.isoformat()

        if not data.get("date_from"):
            data["date_from"] = default_from
        if not data.get("date_to"):
            data["date_to"] = default_to

        form = LateArrivalReportFilterForm(data)
        qs = LateArrival.objects.select_related("student", "responsible").all()

        # 2) Filtramos por rango (siempre válido porque seteamos defaults)
        if form.is_valid():
            d1 = form.cleaned_data["date_from"]
            d2 = form.cleaned_data["date_to"]
            qs = qs.filter(reported_at__date__gte=d1, reported_at__date__lte=d2)
            qs = scope_late_arrivals_for(request.user, qs).order_by("-reported_at")

            # 3) Exportar a Excel si corresponde
            if request.GET.get("export") == "1":
                return self._export_excel(qs, d1, d2)
        else:
            # fallback imposible en práctica, pero por seguridad
            qs = qs.none()
            d1, d2 = default_from, default_to

        # 4) Render normal con form (ya “bound” con defaults) y filas
        context = self.get_context_data(form=form, rows=qs)
        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        # NO devolver HttpResponse aquí; solo dict
        context = super().get_context_data(**kwargs)
        # `form` y `rows` ya vienen desde get(); si no están, setear defaults
        context.setdefault("form", LateArrivalReportFilterForm())
        context.setdefault("rows", LateArrival.objects.none())
        return context

    def _export_excel(self, qs, d1, d2):
        wb = Workbook()
        ws = wb.active
        ws.title = "Llegadas tarde"

        headers = ["Alumno", "Grado", "Fecha y hora", "Motivo"]
        ws.append(headers)

        for n in qs:
            alumno = f"{n.student.last_name}, {n.student.first_name}"
            grado = f"{n.student.level} {n.student.grade}"
            fecha_hora = n.reported_at.strftime("%d/%m/%Y %H:%M")
            ws.append([alumno, grado, fecha_hora, n.reason])

        # Anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 28

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"llegadas_tarde_{d1.isoformat()}_a_{d2.isoformat()}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


# ====== REPORTE TOTALIZADO (solo pantalla, con búsqueda y link) ======
class LateArrivalAggregatedView(LoginRequiredMixin, TemplateView):
    template_name = "avisos/reports_aggregated.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        form = LateArrivalAggregatedFilterForm(self.request.GET or None)

        base_qs = LateArrival.objects.select_related("student")
        base_qs = scope_late_arrivals_for(self.request.user, base_qs)

        # filtro por nombre/apellido (%like%)
        qtext = ""
        if form.is_valid():
            qtext = (form.cleaned_data.get("q") or "").strip()
            if qtext:
                base_qs = base_qs.filter(
                    Q(student__last_name__icontains=qtext) | Q(student__first_name__icontains=qtext)
                )

        last30 = timezone.now() - timedelta(days=30)

        # agregación por alumno+grado
        agg = (base_qs.values("student_id", "student__last_name", "student__first_name",
                              "student__level", "student__grade")
               .annotate(
                   total=Count("id"),
                   last30=Count("id", filter=Q(reported_at__gte=last30))
               )
               .order_by("student__last_name", "student__first_name"))

        ctx["form"] = form
        ctx["rows"] = agg  # lista de dicts
        return ctx
